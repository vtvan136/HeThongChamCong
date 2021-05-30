import os
import shutil
from datetime import date, datetime
from openpyxl import workbook
import pandas as pd
import openpyxl
import SQLite
from openpyxl.styles import Border, Side, Alignment

def getDate():
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y")
    return dt_string.split("/")[0]

def getMonth():
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y")
    return dt_string.split("/")[1]

def getTime():
    now = datetime.now()
    dt_string = now.strftime("%H:%M:%S")
    return dt_string

def checkExistFile(month):
    files = os.listdir("TimeKeeper")
    if month + ".xlsx" in files:
        return True
    else:
        return False

def createFile(month):
    path = "TimeKeeper/" + month+ ".xlsx"
    shutil.copy2("TimeKeeper/Format.xlsx", path)
    work_book = openpyxl.load_workbook(path)
    sheet = work_book.active
    sheet["C1"] = month
    
    users = SQLite.getListUser()
    flag = 3
    for user in users:
        if flag % 2 == 1:
            address1 = "A" + str(flag) #ID
            address2 = "B" + str(flag) #Name
            sheet[address1] = user.id
            sheet[address2] = user.name
            flag = flag + 1
        if flag % 2 == 0:
            address1 = "A" + str(flag) 
            address2 = "B" + str(flag) 
            sheet.merge_cells(address1 +":"+ address2)
            sheet[address1] = "OT"
            flag = flag + 1

    work_book.save(path)

def getAddressOfIDs(path):
    result = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    A = sheet["A"]
    for row in A:
        if isinstance(row.value, int):
            result.append((row.value, (str(row).split(".")[1])[:-1]))
    return result

def getAddressOfDate(path, date):
    result = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row = list(sheet.rows)[1]
    for cell in row:
        if isinstance(cell.value, int) and cell.value == int(date):
            return (str(cell).split(".")[1])[:-2]

    return -1

def checkExistValueInCell(address, path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    if sheet[address] == None:
        return False # cell rong~
    else:
        return True 

def loader():
    month = getMonth()
    date = getDate()
    path = "TimeKeeper/" +month+ ".xlsx"
    address_of_date = getAddressOfDate(path, date)
    address_of_ids = getAddressOfIDs(path)
    return month, date, address_of_date, address_of_ids

def timeKeeper(id, address_of_date, address_of_ids):
    date = getDate()
    month = getMonth()
    path = "TimeKeeper/" +month+ ".xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for add in address_of_ids:
        id_in, address_of_id = add
        if id_in == id:
            break
    address_of_id = address_of_id[1:]

    address = address_of_date + ":" + address_of_id
    if checkExistValueInCell(address, path):
        colint = openpyxl.utils.column_index_from_string(address_of_date)
        nextcell = sheet.cell(row=address_of_id, column=colint+1)
        nextcell.value = "FLAG"
    time = getTime()
    

month, date, address_of_date, address_of_ids = loader()

timeKeeper(id, address_of_date, address_of_ids)


