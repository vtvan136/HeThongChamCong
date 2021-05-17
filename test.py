

from xlsxwriter import workbook, worksheet


def saveNewUserToData(id, name, position, type, password):
    import openpyxl
    from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
    center_aligned_text = Alignment(horizontal="center")
    #Lưu vào sheet User
    workbook = openpyxl.load_workbook(filename="HeThongChamCong\Data\list_user.xlsx")
    sheet = workbook.worksheets[0]
    row = len(sheet["A"]) + 1
    sheet.cell(row, 1).value = id
    sheet.cell(row, 2).value = name
    sheet.cell(row, 3).value = position
    sheet.cell(row, 4).value = type     
    sheet.cell(row, 5).value = password
    #format lại data    
    for i in range(1, 6):
        sheet.cell(row, i).alignment = center_aligned_text

    #Lưu vào sheet Admin or Staff
    if type == "Admin":
        shit = workbook.worksheets[1]
    else:
        shit = workbook.worksheets[2]
    row = len(shit["A"]) + 1
    print(row)
    shit.cell(row, 1).value = id
    shit.cell(row, 2).value = name
    shit.cell(row, 3).value = position
    shit.cell(row, 4).value = password
    #format lại data
    for i in range(1, 5):
        shit.cell(row, i).alignment = center_aligned_text

    workbook.save("HeThongChamCong\Data\list_user.xlsx")

def getNewID():
        '''
        Hệ thống sẽ gợi ý ID là số thứ tự tăng dần
        Phương thức này sẽ lấy ID tiếp theo
        '''
        import openpyxl
        workbook = openpyxl.load_workbook(filename="HeThongChamCong\Data\list_user.xlsx")
        sheet = workbook.worksheets[0]
        return len(sheet["A"]) 

def findUser( id):
    from list_user import getUser
    lst_user = getUser()
    for user in lst_user:
        if user.id == id:
            return user
    return -1

print(findUser(1))